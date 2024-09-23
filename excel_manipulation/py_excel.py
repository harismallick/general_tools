from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random

def conditional_color(path):

    wb = load_workbook(path)
    ws = wb.active
    color_index_limit = 200 # Different color every 400 numbers in cells
    color = rgb_hex_generator()

    for row in ws.iter_rows():
        try:

            if row[3].value == "forward":
                row[3].fill = PatternFill("solid", fgColor="008000")
            elif row[3].value == "reverse":
                row[3].fill = PatternFill("solid", fgColor="FF0000")

            cell_4_value = row[4].value
            # print(type(color_index_limit), type(row[4].value))
            if cell_4_value > color_index_limit:
                # print("this is fine")
                new_color = rgb_hex_generator()
                color = new_color
                color_index_limit += 200
                row[4].fill = PatternFill("solid", fgColor=color)

            elif cell_4_value <= color_index_limit:
                row[4].fill = PatternFill("solid", fgColor=color)

        except(TypeError):
            continue

    wb.save("modified.xlsx")

    return 0

def rgb_hex_generator():
    return f"{random.randint(0, 0xFFFFFF):06x}"

def main():
    path = "test_sheet.xlsx"
    conditional_color(path)

    return 0

if __name__ == '__main__':
    main()